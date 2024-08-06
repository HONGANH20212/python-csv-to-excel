from flask import Flask, request, render_template, redirect, url_for, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
import re
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
import os

app = Flask(__name__)
UPLOAD_FOLDER = 'static/uploads/'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/')
def upload_form():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No file part"
    
    file = request.files['file']
    if file.filename == '':
        return "No selected file"
    
    if file:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)
        output_file = process_file(file_path)
        return send_file(output_file, as_attachment=True)

def text_in_quotes(cell_value):
    pattern = r'\(([^)]+)\)'
    return re.finditer(pattern, cell_value)
def get_env(row):
    value = row["Group hierarchy"]
    list_of_value = value.split(">")
    return list_of_value[1].strip()

def process_file(path):
    temp = 0
    template_path = "D:\\python-csv-excel\\template estimate.xlsx"
    template2_path = "D:\\python-csv-excel\\template-estimate-2.xlsx"
    draft = path.replace('.csv', '_output.xlsx')

    df = pd.read_csv(path, skiprows=7)
    df = df.dropna()
    New_df = pd.DataFrame()
    New_df["Env"] = df.apply(lambda row: get_env(row), axis=1)
    New_df["Service"] = df["Service"]
    New_df["Component"] = df["Description"]
    New_df["Spec"] = df["Configuration summary"]
    New_df["Unit"] = "item"
    New_df["Explain"] = df["Description"]
    New_df["Qty"] = ""
    if ((df["Upfront"] != 0).any()): 
        New_df["Upfront"]=df["Upfront"]
        temp = 1
    else: 
        temp = 2 
    New_df["Monthly"] = df["Monthly"]
    New_df["First 12 months total"] = df["First 12 months total"]
    New_df["Service price"] = ""
    filename = path.replace('.csv', '.xlsx')
    New_df.to_excel(filename, index=False)

    source_wb = load_workbook(filename)
    source_ws = source_wb.active

    if (temp == 1):
        template_wb = load_workbook(template_path)
    else: template_wb = load_workbook(template2_path)
    template_ws = template_wb["Sheet1"]

    for idx, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=9):
        for col_idx, cell_value in enumerate(row, start=1):
            template_ws.cell(row=idx, column=col_idx, value=cell_value)

    template_wb.save(draft)
    wb = load_workbook(draft)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=10):
        for cell in row:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    for row in ws.iter_rows(min_row=9, max_col=4, values_only=False):
        cell = row[3]
        cell_value = cell.value
        if cell_value:
            rich_text_parts = []
            last_end = 0

            for match in text_in_quotes(cell_value):    
                start, end = match.span()
                if start > last_end:
                    rich_text_parts.append(cell_value[last_end:start])
                rich_text_parts.append(TextBlock(InlineFont(color="0099FF"), match.group(0)))
                last_end = end + 1

            if last_end < len(cell_value):
                rich_text_parts.append(cell_value[last_end:])
            rich_text = CellRichText(rich_text_parts)
            cell.value = rich_text
    wb.save(draft)
    return draft

if __name__ == '__main__':
    app.run(debug=True)
